"""
SourceLens — Excel Report Generator
Creates a multi-sheet Excel audit report using openpyxl.
"""

import logging
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from app.models import Project, MetricRecord, VerificationStatus

logger = logging.getLogger(__name__)

# Style constants
HEADER_FONT = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="1a1a2e", end_color="1a1a2e", fill_type="solid")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
THIN_BORDER = Border(
    left=Side(style="thin", color="E0E0E0"),
    right=Side(style="thin", color="E0E0E0"),
    top=Side(style="thin", color="E0E0E0"),
    bottom=Side(style="thin", color="E0E0E0"),
)

STATUS_FILLS = {
    VerificationStatus.VERIFIED: PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
    VerificationStatus.PARTIALLY_VERIFIED: PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
    VerificationStatus.UNVERIFIED: PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
    VerificationStatus.CONTRADICTED: PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid"),
    VerificationStatus.NOT_FOUND: PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
}


def generate_excel_report(project: Project, metrics: list[MetricRecord], output_path: str):
    """Generate a comprehensive Excel audit report."""
    wb = Workbook()

    # Sheet 1: Summary
    _create_summary_sheet(wb, project, metrics)

    # Sheet 2: Metric Detail
    _create_metric_detail_sheet(wb, metrics)

    # Sheet 3: Source Document Index
    _create_source_index_sheet(wb, project)

    # Sheet 4: Citation Chains
    _create_citation_chain_sheet(wb, metrics)

    # Sheet 5: Issues & Flags
    _create_issues_sheet(wb, metrics)

    wb.save(output_path)
    logger.info(f"Excel report saved to {output_path}")


def _style_header_row(ws, num_cols: int):
    """Apply header styling to the first row."""
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER


def _auto_width(ws, min_width: int = 10, max_width: int = 50):
    """Auto-adjust column widths."""
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except Exception:
                pass
        adjusted = min(max(max_length + 2, min_width), max_width)
        ws.column_dimensions[col_letter].width = adjusted


def _create_summary_sheet(wb: Workbook, project: Project, metrics: list[MetricRecord]):
    """Create the Summary sheet."""
    ws = wb.active
    ws.title = "Summary"

    # Title
    ws["A1"] = "SourceLens Audit Report"
    ws["A1"].font = Font(name="Calibri", bold=True, size=16, color="1a1a2e")
    ws["A2"] = f"Project: {project.project_name}"
    ws["A3"] = f"Generated: {project.created_at.strftime('%Y-%m-%d %H:%M')}"
    ws["A4"] = f"Presentation: {project.presentation_file.original_filename if project.presentation_file else 'N/A'}"

    # Summary stats
    row = 6
    ws.cell(row=row, column=1, value="Audit Summary").font = Font(bold=True, size=13)
    row += 1

    summary = project.summary_stats
    if summary:
        stats = [
            ("Total Metrics Extracted", summary.total_metrics_extracted),
            ("Verified", summary.verified),
            ("Partially Verified", summary.partially_verified),
            ("Unverified", summary.unverified),
            ("Contradicted", summary.contradicted),
            ("Not Found", summary.not_found),
            ("Average Confidence", f"{summary.avg_confidence_score:.1%}"),
        ]
        for label, value in stats:
            ws.cell(row=row, column=1, value=label)
            ws.cell(row=row, column=2, value=value)
            row += 1

        # Flags
        if summary.flags:
            row += 1
            ws.cell(row=row, column=1, value="Flags & Warnings").font = Font(bold=True, size=13)
            row += 1
            for flag in summary.flags:
                ws.cell(row=row, column=1, value=flag)
                row += 1

    _auto_width(ws)


def _create_metric_detail_sheet(wb: Workbook, metrics: list[MetricRecord]):
    """Create the Metric Detail sheet — one row per metric."""
    ws = wb.create_sheet("Metric Detail")

    headers = [
        "Slide #", "Metric Value", "Type", "Description", "Context",
        "Status", "Confidence", "Source Document", "Page", "Section",
        "Exact Quote", "Ultimate Source", "Source Type", "Reliability",
        "Stars", "Notes", "LLM Reasoning",
    ]
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
    _style_header_row(ws, len(headers))

    for r, m in enumerate(metrics, 2):
        v = m.verification
        ws.cell(row=r, column=1, value=m.slide_number)
        ws.cell(row=r, column=2, value=m.extracted_metric.value)
        ws.cell(row=r, column=3, value=m.extracted_metric.metric_type.value)
        ws.cell(row=r, column=4, value=m.extracted_metric.description)
        ws.cell(row=r, column=5, value=m.extracted_metric.context_sentence[:200])

        if v:
            status_cell = ws.cell(row=r, column=6, value=v.status.value.upper())
            status_cell.fill = STATUS_FILLS.get(v.status, PatternFill())
            ws.cell(row=r, column=7, value=f"{v.confidence_score:.0%}")
            ws.cell(row=r, column=8, value=v.matched_source_document_name or "")
            ws.cell(row=r, column=9, value=v.matched_location.page_number if v.matched_location else "")
            ws.cell(row=r, column=10, value=v.matched_location.section_heading if v.matched_location else "")
            ws.cell(row=r, column=11, value=(v.matched_location.exact_quote[:300] if v.matched_location else ""))
            ws.cell(row=r, column=12, value=v.citation_chain.ultimate_source if v.citation_chain else "")
            ws.cell(row=r, column=13, value=v.citation_chain.ultimate_source_type if v.citation_chain else "")
            ws.cell(row=r, column=14, value=v.source_reliability.tier_label if v.source_reliability else "")
            ws.cell(row=r, column=15, value=v.source_reliability.stars if v.source_reliability else "")
            ws.cell(row=r, column=16, value=v.verification_notes[:200] if v.verification_notes else "")
            ws.cell(row=r, column=17, value=v.llm_reasoning[:300] if v.llm_reasoning else "")

        # Apply border
        for c in range(1, len(headers) + 1):
            ws.cell(row=r, column=c).border = THIN_BORDER

    _auto_width(ws)


def _create_source_index_sheet(wb: Workbook, project: Project):
    """Create the Source Document Index sheet."""
    ws = wb.create_sheet("Source Documents")
    headers = ["Filename", "Label", "Type", "Pages", "Chunks", "Bibliography Entries", "File Size (KB)"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
    _style_header_row(ws, len(headers))

    for r, doc in enumerate(project.source_documents, 2):
        ws.cell(row=r, column=1, value=doc.original_filename)
        ws.cell(row=r, column=2, value=doc.document_label)
        ws.cell(row=r, column=3, value=doc.document_type.value)
        ws.cell(row=r, column=4, value=doc.page_count)
        ws.cell(row=r, column=5, value=doc.chunk_count)
        ws.cell(row=r, column=6, value=len(doc.bibliography_entries))
        ws.cell(row=r, column=7, value=round(doc.file_size_bytes / 1024, 1))
    _auto_width(ws)


def _create_citation_chain_sheet(wb: Workbook, metrics: list[MetricRecord]):
    """Create the Citation Chain sheet."""
    ws = wb.create_sheet("Citation Chains")
    headers = ["Slide #", "Metric", "Intermediate Source", "Ref ID",
               "Ultimate Source", "Source Type", "Reliability Tier", "Stars", "Recommendation"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
    _style_header_row(ws, len(headers))

    r = 2
    for m in metrics:
        v = m.verification
        if v and v.citation_chain and v.citation_chain.ultimate_source:
            cc = v.citation_chain
            ws.cell(row=r, column=1, value=m.slide_number)
            ws.cell(row=r, column=2, value=m.extracted_metric.value)
            ws.cell(row=r, column=3, value=cc.intermediate_source or "")
            ws.cell(row=r, column=4, value=cc.intermediate_ref_id or "")
            ws.cell(row=r, column=5, value=cc.ultimate_source)
            ws.cell(row=r, column=6, value=cc.ultimate_source_type or "")
            if v.source_reliability:
                ws.cell(row=r, column=7, value=v.source_reliability.tier_label)
                ws.cell(row=r, column=8, value=v.source_reliability.stars)
                ws.cell(row=r, column=9, value=v.source_reliability.recommendation)
            r += 1
    _auto_width(ws)


def _create_issues_sheet(wb: Workbook, metrics: list[MetricRecord]):
    """Create the Issues & Flags sheet — only metrics with problems."""
    ws = wb.create_sheet("Issues & Flags")
    headers = ["Severity", "Slide #", "Metric", "Status", "Confidence", "Issue Description"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
    _style_header_row(ws, len(headers))

    r = 2
    for m in metrics:
        v = m.verification
        if v and v.status != VerificationStatus.VERIFIED:
            severity = {
                VerificationStatus.CONTRADICTED: "🔴 CRITICAL",
                VerificationStatus.UNVERIFIED: "🟠 HIGH",
                VerificationStatus.NOT_FOUND: "🟡 MEDIUM",
                VerificationStatus.PARTIALLY_VERIFIED: "🟢 LOW",
            }.get(v.status, "⚪ INFO")

            ws.cell(row=r, column=1, value=severity)
            ws.cell(row=r, column=2, value=m.slide_number)
            ws.cell(row=r, column=3, value=m.extracted_metric.value)
            ws.cell(row=r, column=4, value=v.status.value.upper())
            ws.cell(row=r, column=5, value=f"{v.confidence_score:.0%}")
            ws.cell(row=r, column=6, value=v.verification_notes or v.llm_reasoning[:200])
            r += 1
    _auto_width(ws)
